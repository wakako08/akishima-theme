#!/usr/bin/env python3
"""
掲載事項調査.xlsx → assets/data/communities/{slug}.json

- スラッグは communities-data.php の会員番号で確定（マスター優先）
- JSON の name はマスター正式名称を使用
- 自由記述（列16）:
    原則 → facility_rental
    例外 INTRO_FREE_TEXT_NOS → intro 追記

Usage:
  python3 scripts/import-community-survey.py --xlsx "/path/to.xlsx"
  python3 scripts/import-community-survey.py --xlsx "/path/to.xlsx" --only 1,2,3
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print('openpyxl が必要です: pip install openpyxl', file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
OUT_JSON = ROOT / 'assets' / 'data' / 'communities'
MASTER_PHP = ROOT / 'inc' / 'communities-data.php'

# 自由記述を紹介文へ回す会員番号（施設説明ではない例外）
INTRO_FREE_TEXT_NOS = {8}  # 東町中央自治会


def norm(value) -> str:
    if value is None:
        return ''
    text = str(value).replace('\r\n', '\n').replace('\r', '\n')
    return '\n'.join(line.rstrip() for line in text.split('\n')).strip()


def normalize_name(name: str) -> str:
    """比較用に名前を正規化（半角カナ→全角、空白除去など）"""
    name = unicodedata.normalize('NFKC', name or '')
    name = name.replace(' ', '').replace('　', '')
    name = name.replace('ヶ', 'ケ').replace('ノ', 'の')
    return name


def load_master() -> dict[int, dict]:
    """
    member_no -> {block, slug, name}
    akishima_get_blocks_data_raw() 内の members をパース
    """
    php = MASTER_PHP.read_text(encoding='utf-8')
    # raw 関数の中身だけ
    m = re.search(
        r'function akishima_get_blocks_data_raw\(\)\s*\{\s*return array\((.*)\n\s*\);\s*\}',
        php,
        re.S,
    )
    if not m:
        raise RuntimeError('akishima_get_blocks_data_raw() をパースできません')

    body = m.group(1)
    master: dict[int, dict] = {}
    block_no = None
    pending_block_no = None

    for line in body.splitlines():
        # ブロック定義は複数行:
        #   'no'   => 1,
        #   'name' => '第1ブロック',
        if pending_block_no is None:
            bm = re.match(r"\s*'no'\s*=>\s*(\d+)\s*,\s*$", line)
            if bm:
                pending_block_no = int(bm.group(1))
                continue
        else:
            nm = re.search(r"'name'\s*=>\s*'(第\d+ブロック)'", line)
            if nm:
                block_no = pending_block_no
                pending_block_no = None
                continue
            # メンバー行の 'no' と誤認しないよう、次行がブロック名でなければ破棄
            if re.search(r"array\(\s*'no'\s*=>", line) or "'members'" in line:
                pending_block_no = None

        mm = re.search(r"array\(\s*'no'\s*=>\s*(\d+)\s*,\s*'name'\s*=>\s*'([^']+)'", line)
        if mm and block_no is not None:
            mno = int(mm.group(1))
            name = mm.group(2)
            if mno in master:
                raise RuntimeError(f'会員番号の重複: {mno}')
            master[mno] = {
                'block': block_no,
                'slug': f'{block_no:02d}-{mno:02d}',
                'name': name,
            }
    if not master:
        raise RuntimeError('マスター会員が0件です（パース失敗）')
    return master


def compose_facility_rental(rental_yn: str, note: str) -> str:
    parts = []
    if rental_yn:
        parts.append(rental_yn)
    if note:
        parts.append(note)
    return '\n'.join(parts)


def append_intro(intro: str, note: str) -> str:
    intro = norm(intro)
    note = norm(note)
    if not note:
        return intro
    if not intro:
        return note
    if note in intro:
        return intro
    return intro + '\n\n' + note


def classify_free_text(member_no: int, note: str) -> str:
    if not note:
        return 'facility_rental'
    if member_no in INTRO_FREE_TEXT_NOS:
        return 'intro'
    facility_markers = (
        '貸', '使用', '部屋', '会館', '集会', '施設', '倶楽部', '公会堂', 'サロン', '料金',
    )
    if any(m in note for m in facility_markers):
        return 'facility_rental'
    if len(note) > 80:
        return 'intro'
    return 'facility_rental'


def load_rows(xlsx: Path, only_nos: set[int] | None, master: dict[int, dict]):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb['最新内容']
    current_block = None
    rows = []
    warnings = []

    for r in range(3, ws.max_row + 1):
        block_cell = ws.cell(r, 1).value
        if block_cell is not None and str(block_cell).strip() != '':
            try:
                current_block = int(block_cell)
            except ValueError:
                pass

        no = ws.cell(r, 2).value
        ss_name = ws.cell(r, 3).value
        if no is None or not ss_name:
            continue
        try:
            no = int(no)
        except ValueError:
            warnings.append(f'R{r}: 会員番号が数値でない: {no!r}')
            continue
        if only_nos is not None and no not in only_nos:
            continue

        ss_name = str(ss_name).strip()
        if no not in master:
            warnings.append(f'R{r}: マスターに無い会員番号 {no} ({ss_name}) → スキップ')
            continue

        info = master[no]
        slug = info['slug']
        official_name = info['name']

        if current_block is not None and current_block != info['block']:
            warnings.append(
                f'R{r}: ブロック不一致 №{no} スプシ={current_block} マスター={info["block"]} '
                f'→ マスターの {slug} を採用'
            )

        if normalize_name(ss_name) != normalize_name(official_name):
            warnings.append(
                f'R{r}: 名称差 №{no} スプシ={ss_name!r} マスター={official_name!r} '
                f'→ JSON name はマスターを使用'
            )

        intro = norm(ws.cell(r, 8).value)
        activities = norm(ws.cell(r, 9).value)
        fee = norm(ws.cell(r, 10).value)
        organizations = norm(ws.cell(r, 11).value)
        facility_name = norm(ws.cell(r, 13).value)
        facility_address = norm(ws.cell(r, 14).value)
        rental_yn = norm(ws.cell(r, 15).value)
        free_note = norm(ws.cell(r, 16).value)

        # 「未記入」は空扱いに近いが、提出値として残す（管理者が後で直せる）
        route = classify_free_text(no, free_note)
        facility_rental = compose_facility_rental(
            rental_yn,
            free_note if route == 'facility_rental' else '',
        )
        if route == 'intro' and free_note:
            intro = append_intro(intro, free_note)

        has_text = any(
            [
                intro,
                activities,
                fee,
                organizations,
                facility_name,
                facility_address,
                facility_rental,
            ]
        )
        if not has_text:
            continue

        data = {
            'name': official_name,
            'intro': intro,
            'activities': activities,
            'fee': fee,
            'organizations': organizations,
            'facility_name': facility_name,
            'facility_address': facility_address,
            'facility_rental': facility_rental,
            'events': [],
        }
        rows.append(
            {
                'slug': slug,
                'no': no,
                'ss_name': ss_name,
                'route': route if free_note else None,
                'data': data,
            }
        )

    return rows, warnings


def main() -> int:
    parser = argparse.ArgumentParser(description='掲載事項調査 → communities JSON')
    parser.add_argument('--xlsx', required=True, help='調査xlsxのパス')
    parser.add_argument('--only', default='', help='対象会員番号（カンマ区切り）')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.is_file():
        print(f'xlsx が見つかりません: {xlsx}', file=sys.stderr)
        return 1

    only_nos = None
    if args.only.strip():
        only_nos = {int(x.strip()) for x in args.only.split(',') if x.strip()}

    master = load_master()
    print(f'master members: {len(master)}')

    rows, warnings = load_rows(xlsx, only_nos, master)
    for w in warnings:
        print(f'WARN: {w}')

    # 重複スラッグチェック
    slugs = [r['slug'] for r in rows]
    if len(slugs) != len(set(slugs)):
        print('ERROR: スラッグ重複', file=sys.stderr)
        return 1

    OUT_JSON.mkdir(parents=True, exist_ok=True)
    by_block: dict[int, int] = {}

    for row in rows:
        path = OUT_JSON / f'{row["slug"]}.json'
        route = row['route']
        route_label = f' free→{route}' if route else ''
        block = int(row['slug'].split('-')[0])
        by_block[block] = by_block.get(block, 0) + 1
        if args.dry_run:
            print(
                f'[dry-run] {path.name} {row["data"]["name"]} '
                f'fee={row["data"]["fee"]!r}{route_label}'
            )
            continue
        path.write_text(
            json.dumps(row['data'], ensure_ascii=False, indent=2) + '\n',
            encoding='utf-8',
        )
        print(f'wrote {path.relative_to(ROOT)} ({row["data"]["name"]}){route_label}')

    print('---')
    print(f'done: {len(rows)} files')
    for b in sorted(by_block):
        print(f'  block {b:02d}: {by_block[b]}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
